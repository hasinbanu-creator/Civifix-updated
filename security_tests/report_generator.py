"""Security vulnerability report generator"""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from security_scanner import scan_civifix_security


class VulnerabilityReportGenerator:
    """Generate vulnerability reports"""
    
    def __init__(self):
        self.findings = scan_civifix_security()
        self.severity_colors = {
            "CRITICAL": "FF0000",
            "HIGH": "FFC7CE",
            "MEDIUM": "FFEB9C",
            "LOW": "92D050"
        }
    
    def generate_excel_report(self, filename="vulnerability_report.xlsx"):
        """Generate Excel vulnerability report"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Vulnerabilities"
        
        # Define styles
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Headers
        headers = [
            "Severity",
            "File Path",
            "Vulnerability",
            "Description",
            "Remediation",
            "Line #",
            "CWE",
            "CVSS Score"
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        
        # Add findings
        row_num = 2
        for finding in self.findings:
            severity = finding.get("severity", "LOW")
            
            ws.cell(row=row_num, column=1).value = severity
            ws.cell(row=row_num, column=2).value = finding.get("file", "")
            ws.cell(row=row_num, column=3).value = finding.get("vulnerability", "")
            ws.cell(row=row_num, column=4).value = finding.get("description", "")
            ws.cell(row=row_num, column=5).value = finding.get("remediation", "")
            ws.cell(row=row_num, column=6).value = finding.get("line", 1)
            ws.cell(row=row_num, column=7).value = self._get_cwe(finding.get("vulnerability"))
            ws.cell(row=row_num, column=8).value = self._get_cvss_score(severity)
            
            # Color code by severity
            severity_color = self.severity_colors.get(severity, "FFFFFF")
            severity_fill = PatternFill(start_color=severity_color, end_color=severity_color, fill_type="solid")
            
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if col == 1:  # Severity column
                    cell.fill = severity_fill
            
            row_num += 1
        
        # Adjust column widths
        column_widths = [12, 30, 25, 35, 35, 8, 10, 12]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col_num)].width = width
        
        # Add summary sheet
        self._add_summary_sheet(wb)
        
        # Add remediation sheet
        self._add_remediation_sheet(wb)
        
        wb.save(filename)
        print(f"Report saved: {filename}")
        return filename
    
    def _add_summary_sheet(self, wb):
        """Add summary sheet"""
        summary_ws = wb.create_sheet("Summary", 1)
        
        severity_counts = {}
        for finding in self.findings:
            severity = finding.get("severity", "LOW")
            severity_counts[severity] = severity_counts.get(severity, 0) + 1
        
        total_findings = len(self.findings)
        
        summary_data = [
            ("Total Vulnerabilities", total_findings),
            ("Critical", severity_counts.get("CRITICAL", 0)),
            ("High", severity_counts.get("HIGH", 0)),
            ("Medium", severity_counts.get("MEDIUM", 0)),
            ("Low", severity_counts.get("LOW", 0)),
            ("", ""),
            ("Scan Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Scan Type", "Source Code Analysis"),
            ("Scope", "CiviFix Backend Application"),
        ]
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for row_num, (label, value) in enumerate(summary_data, 1):
            label_cell = summary_ws.cell(row=row_num, column=1)
            value_cell = summary_ws.cell(row=row_num, column=2)
            
            label_cell.value = label
            value_cell.value = value
            
            if row_num <= 5:  # Highlight severity counts
                label_cell.fill = header_fill
                label_cell.font = header_font
            
            summary_ws.column_dimensions['A'].width = 25
            summary_ws.column_dimensions['B'].width = 25
    
    def _add_remediation_sheet(self, wb):
        """Add remediation recommendations sheet"""
        remediation_ws = wb.create_sheet("Remediation Plan", 2)
        
        headers = ["Priority", "Vulnerability", "Recommended Action", "Timeline", "Owner"]
        
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col_num, header in enumerate(headers, 1):
            cell = remediation_ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Add remediation items
        row_num = 2
        critical_remediation = [
            ("IMMEDIATE", "Hardcoded Secrets", "Move all secrets to environment variables or AWS Secrets Manager", "1-3 days", "Backend Team"),
            ("IMMEDIATE", "Code Injection", "Remove exec/eval usage; use safe alternatives", "1-2 days", "Backend Team"),
            ("URGENT", "Insecure CORS", "Restrict CORS to specific trusted domains", "1 day", "Backend Team"),
            ("URGENT", "SQL/NoSQL Injection", "Implement parameterized queries; use ORM", "2-5 days", "Backend Team"),
            ("HIGH", "Missing Input Validation", "Add validators for all inputs", "3-5 days", "Backend Team"),
            ("HIGH", "Role-based Access Control", "Implement centralized RBAC", "3-7 days", "Backend Team"),
            ("MEDIUM", "Path Traversal", "Validate file paths; use secure operations", "2-3 days", "Backend Team"),
            ("MEDIUM", "Vulnerable Dependencies", "Update packages; use pip-audit", "1-2 days", "DevOps Team"),
        ]
        
        for priority, vuln, action, timeline, owner in critical_remediation:
            remediation_ws.cell(row=row_num, column=1).value = priority
            remediation_ws.cell(row=row_num, column=2).value = vuln
            remediation_ws.cell(row=row_num, column=3).value = action
            remediation_ws.cell(row=row_num, column=4).value = timeline
            remediation_ws.cell(row=row_num, column=5).value = owner
            row_num += 1
        
        # Adjust column widths
        remediation_ws.column_dimensions['A'].width = 12
        remediation_ws.column_dimensions['B'].width = 25
        remediation_ws.column_dimensions['C'].width = 40
        remediation_ws.column_dimensions['D'].width = 15
        remediation_ws.column_dimensions['E'].width = 15
    
    def _get_cwe(self, vulnerability):
        """Get CWE mapping"""
        cwe_mapping = {
            "Hardcoded Secrets": "CWE-798",
            "Code Injection": "CWE-94",
            "SQL Injection": "CWE-89",
            "NoSQL Injection": "CWE-943",
            "Path Traversal": "CWE-22",
            "Missing Authentication": "CWE-306",
            "Missing Authorization": "CWE-862",
            "Insecure CORS Configuration": "CWE-346",
            "Command Injection": "CWE-78",
        }
        return cwe_mapping.get(vulnerability, "N/A")
    
    def _get_cvss_score(self, severity):
        """Get CVSS score based on severity"""
        cvss_mapping = {
            "CRITICAL": 9.9,
            "HIGH": 7.5,
            "MEDIUM": 5.3,
            "LOW": 3.3
        }
        return cvss_mapping.get(severity, 0.0)
    
    def generate_markdown_report(self, filename="vulnerability_analysis.md"):
        """Generate Markdown report"""
        findings_by_severity = {}
        for finding in self.findings:
            severity = finding.get("severity")
            if severity not in findings_by_severity:
                findings_by_severity[severity] = []
            findings_by_severity[severity].append(finding)
        
        content = []
        content.append("# Security Vulnerability Assessment Report\n")
        content.append(f"**Assessment Date:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        content.append(f"**Application:** CiviFix\n")
        content.append(f"**Module:** Backend API\n")
        content.append(f"**Total Vulnerabilities:** {len(self.findings)}\n\n")
        
        # Summary
        content.append("## Executive Summary\n")
        for severity in ["CRITICAL", "HIGH", "MEDIUM", "LOW"]:
            count = len(findings_by_severity.get(severity, []))
            if count > 0:
                content.append(f"- **{severity}:** {count} vulnerabilities\n")
        
        content.append("\n")
        
        # Detailed findings
        for severity in ["CRITICAL", "HIGH", "MEDIUM", "LOW"]:
            if severity not in findings_by_severity:
                continue
            
            findings = findings_by_severity[severity]
            content.append(f"## {severity} Severity Vulnerabilities\n\n")
            
            for finding in findings:
                content.append(f"### {finding['vulnerability']}\n\n")
                content.append(f"**File:** `{finding['file']}`  \n")
                content.append(f"**Line:** {finding['line']}  \n\n")
                content.append(f"**Description:**  \n{finding['description']}\n\n")
                content.append(f"**Remediation:**  \n{finding['remediation']}\n\n")
                content.append("---\n\n")
        
        with open(filename, 'w') as f:
            f.writelines(content)
        
        print(f"Markdown report saved: {filename}")
        return filename


def main():
    """Main execution"""
    print("CiviFix Security Vulnerability Assessment")
    print("=" * 50)
    
    generator = VulnerabilityReportGenerator()
    
    # Generate Excel report
    print("\nGenerating Excel report...")
    generator.generate_excel_report("vulnerability_report.xlsx")
    
    # Generate Markdown report
    print("Generating Markdown report...")
    generator.generate_markdown_report("vulnerability_analysis.md")
    
    print("\n" + "=" * 50)
    print("Reports generated successfully")


if __name__ == "__main__":
    main()
